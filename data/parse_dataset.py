import json
import os
import openpyxl

DATASET_PATH = os.path.join(os.path.dirname(__file__), "Gen_AI_Dataset.xlsx")


def parse():
    wb = openpyxl.load_workbook(DATASET_PATH)

    # ── Train Set ──
    ws = wb["Train-Set"]
    train = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] == "Query":
            continue
        q, url = row
        if q not in train:
            train[q] = []
        train[q].append(url)

    train_list = [{"query": q, "relevant_urls": urls} for q, urls in train.items()]
    with open(os.path.join(os.path.dirname(__file__), "train_data.json"), "w") as f:
        json.dump(train_list, f, indent=2)
    print(f"✅ Saved train_data.json ({len(train_list)} queries)")

    # ── Test Set ──
    ws2 = wb["Test-Set"]
    test = []
    for row in ws2.iter_rows(values_only=True):
        if row[0] == "Query":
            continue
        test.append(row[0])

    with open(os.path.join(os.path.dirname(__file__), "test_queries.json"), "w") as f:
        json.dump(test, f, indent=2)
    print(f"✅ Saved test_queries.json ({len(test)} queries)")


if __name__ == "__main__":
    parse()
